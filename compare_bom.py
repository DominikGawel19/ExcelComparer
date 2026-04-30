import shutil
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

OLD_FILE = '10222-ODS-E-F10-PIP-LST-0001-R04-LISTA MATERIAŁOWA.xlsx'
NEW_FILE = '10222-ODS-E-F10-PIP-LST-0001-R05-Lista materiałowa.xlsx'
OUT_FILE = '10222-ODS-E-F10-PIP-LST-0001-R05-Lista materiałowa_POROWNANIE.xlsx'

YELLOW_FILL = PatternFill(fill_type='solid', start_color='FFFFFF00', end_color='FFFFFF00')
RED_COLOR = 'FFFF0000'

HEADER_ROW = 10   # 1-based Excel row with column headers
DATA_START = 11   # 1-based Excel row where data begins

DATA_SHEETS = ['LM', 'LP', 'LS', 'LW', 'LB']


def vals_equal(a, b):
    a_none = a is None or (isinstance(a, float) and pd.isna(a))
    b_none = b is None or (isinstance(b, float) and pd.isna(b))
    if a_none and b_none:
        return True
    if a_none or b_none:
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        fa, fb = float(a), float(b)
        diff = abs(fa - fb)
        mag = max(abs(fa), abs(fb), 1e-12)
        return diff / mag < 1e-6  # 1ppm relative tolerance
    return str(a).strip() == str(b).strip()


def is_pozycja(val):
    if val is None:
        return False
    return bool(re.match(r'^[A-Z]{1,4}-\d+', str(val).strip()))


def is_parent_ls(vals):
    """LS parent: has position but no profile (col 1) and no material (col 2)."""
    if not is_pozycja(vals[0]):
        return False
    b_empty = len(vals) < 2 or vals[1] is None or str(vals[1]).strip() == ''
    c_empty = len(vals) < 3 or vals[2] is None or str(vals[2]).strip() == ''
    return b_empty and c_empty


def read_data_rows(wb_data_only, sheet_name):
    """
    Read all data rows using data_only workbook.
    Returns list of (excel_row_1based, [cell_values]).
    """
    ws = wb_data_only[sheet_name]
    rows = []
    for excel_row in range(DATA_START, ws.max_row + 1):
        vals = [ws.cell(row=excel_row, column=c).value
                for c in range(1, ws.max_column + 1)]
        rows.append((excel_row, vals))
    return rows


def build_old_map(old_rows, is_ls=False):
    """
    Map: pozycja -> list of (excel_row, vals).
    For LS, stores parent and child entries separately.
    """
    result = {}
    for excel_row, vals in old_rows:
        if not is_pozycja(vals[0]):
            continue
        key = str(vals[0]).strip()
        if is_ls:
            entry_type = 'parent' if is_parent_ls(vals) else 'child'
            composite = (key, entry_type)
        else:
            composite = key
        if composite not in result:
            result[composite] = []
        result[composite].append((excel_row, vals))
    return result


def build_ls_child_to_parent(rows):
    """Return dict: child_pozycja -> parent_pozycja."""
    child_to_parent = {}
    current_parent = None
    for _excel_row, vals in rows:
        if not is_pozycja(vals[0]):
            continue
        if is_parent_ls(vals):
            current_parent = str(vals[0]).strip()
        else:
            if current_parent:
                child_to_parent[str(vals[0]).strip()] = current_parent
    return child_to_parent


def apply_yellow_row(ws, excel_row, max_col):
    for c in range(1, max_col + 1):
        ws.cell(row=excel_row, column=c).fill = YELLOW_FILL


def apply_red_text(ws, excel_row, col_1based):
    cell = ws.cell(row=excel_row, column=col_1based)
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=f.bold, italic=f.italic,
        underline=f.underline, strike=f.strike, vertAlign=f.vertAlign,
        color=RED_COLOR
    )


def compare_sheet(ws_out, old_rows, new_rows, is_ls=False):
    """
    Compare new rows vs old rows and apply formatting to ws_out.
    Returns set of changed pozycja strings (children only for LS).
    """
    max_col = ws_out.max_column
    old_map = build_old_map(old_rows, is_ls=is_ls)
    changed_positions = set()

    for excel_row, new_vals in new_rows:
        if not is_pozycja(new_vals[0]):
            continue

        pozycja = str(new_vals[0]).strip()
        parent = is_ls and is_parent_ls(new_vals)

        if is_ls:
            key = (pozycja, 'parent' if parent else 'child')
        else:
            key = pozycja

        if key not in old_map:
            # New item not in old version
            apply_yellow_row(ws_out, excel_row, max_col)
            changed_positions.add(pozycja)
            continue

        old_vals = old_map[key][0][1]
        n_cols = max(len(new_vals), len(old_vals))

        changed_cols = []
        for i in range(n_cols):
            nv = new_vals[i] if i < len(new_vals) else None
            ov = old_vals[i] if i < len(old_vals) else None
            if not vals_equal(nv, ov):
                changed_cols.append(i + 1)  # 1-based

        if changed_cols:
            apply_yellow_row(ws_out, excel_row, max_col)
            for col in changed_cols:
                apply_red_text(ws_out, excel_row, col)
            changed_positions.add(pozycja)

    return changed_positions


def mark_ls_parents(ws_out, new_rows, changed_children, child_to_parent):
    """Mark parent rows yellow if any of their children changed."""
    parents_to_mark = {child_to_parent[c] for c in changed_children if c in child_to_parent}
    if not parents_to_mark:
        return
    max_col = ws_out.max_column
    for excel_row, vals in new_rows:
        if is_pozycja(vals[0]) and is_parent_ls(vals):
            if str(vals[0]).strip() in parents_to_mark:
                apply_yellow_row(ws_out, excel_row, max_col)


def process_sheet(wb_old_d, wb_new_d, wb_out, sheet_name):
    print(f'Processing {sheet_name}...')
    old_rows = read_data_rows(wb_old_d, sheet_name)
    new_rows = read_data_rows(wb_new_d, sheet_name)
    ws_out = wb_out[sheet_name]

    is_ls = sheet_name == 'LS'
    changed = compare_sheet(ws_out, old_rows, new_rows, is_ls=is_ls)

    if is_ls:
        child_to_parent = build_ls_child_to_parent(new_rows)
        changed_children = {p for p in changed
                            if (p, 'child') in build_old_map(new_rows, is_ls=True) or True}
        # Filter: only non-parent items that changed
        changed_children = set()
        for _r, vals in new_rows:
            if is_pozycja(vals[0]) and not is_parent_ls(vals):
                if str(vals[0]).strip() in changed:
                    changed_children.add(str(vals[0]).strip())
        mark_ls_parents(ws_out, new_rows, changed_children, child_to_parent)

    print(f'  Changed positions: {len(changed)}')


def main():
    shutil.copy2(NEW_FILE, OUT_FILE)

    wb_old_d = load_workbook(OLD_FILE, data_only=True)
    wb_new_d = load_workbook(NEW_FILE, data_only=True)
    wb_out = load_workbook(OUT_FILE)  # no data_only - preserves formulas

    for sheet in DATA_SHEETS:
        if sheet in wb_old_d.sheetnames and sheet in wb_out.sheetnames:
            process_sheet(wb_old_d, wb_new_d, wb_out, sheet)

    wb_out.save(OUT_FILE)
    print(f'\nSaved: {OUT_FILE}')


if __name__ == '__main__':
    main()
