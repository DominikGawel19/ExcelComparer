import os
import re
import shutil
import threading
import tkinter as tk
import zipfile
from io import BytesIO
from tkinter import filedialog, messagebox, ttk
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from tkinterdnd2 import TkinterDnD, DND_FILES

YELLOW_FILL = PatternFill(fill_type='solid', start_color='FFFFFF00', end_color='FFFFFF00')
DELETE_FILL = PatternFill(fill_type='solid', start_color='FFFFC7CE', end_color='FFFFC7CE')
RED_COLOR = 'FFFF0000'
MAX_COL = 9       # Fill yellow only up to column I
OLD_COL_START = 13  # Column M — old values placed here for changed rows
DATA_SHEETS = ['LM', 'LP', 'LS', 'LW', 'LB', 'LG']


# ── Comparison logic ─────────────────────────────────────────────────────────

def vals_equal(a, b):
    def is_empty(v):
        return (v is None
                or (isinstance(v, float) and pd.isna(v))
                or (isinstance(v, str) and v.strip() == ''))
    if is_empty(a) and is_empty(b):
        return True
    if is_empty(a) or is_empty(b):
        return False
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        fa, fb = float(a), float(b)
        diff = abs(fa - fb)
        mag = max(abs(fa), abs(fb), 1e-12)
        return diff / mag < 1e-6
    return str(a).strip() == str(b).strip()


def is_pozycja(val):
    if val is None:
        return False
    return bool(re.match(r'^[A-Z]{1,4}-\d+', str(val).strip()))


def is_lb_data_row(vals):
    """LB data row: any non-empty col A that is not the header line."""
    v = vals[0]
    if v is None:
        return False
    s = str(v).strip()
    return s != '' and s != 'Pozycja:'


def is_parent_ls(vals):
    if not is_pozycja(vals[0]):
        return False
    b_empty = len(vals) < 2 or vals[1] is None or str(vals[1]).strip() == ''
    c_empty = len(vals) < 3 or vals[2] is None or str(vals[2]).strip() == ''
    return b_empty and c_empty


def read_data_rows(wb_data_only, sheet_name, data_start):
    ws = wb_data_only[sheet_name]
    rows = []
    for excel_row in range(data_start, ws.max_row + 1):
        vals = [ws.cell(row=excel_row, column=c).value
                for c in range(1, ws.max_column + 1)]
        rows.append((excel_row, vals))
    return rows


def detect_data_start(ws):
    """
    Scan for a header row (>= 3 string cells containing ':') within the first 50 rows.
    Returns (data_start_row, data_col_end) both 1-based.
    data_col_end = rightmost column with a label in the header row.
    Falls back to (2, ws.max_column) if no header found.
    """
    def is_label(v):
        return v is not None and isinstance(v, str) and ':' in v and v.strip()

    for r in range(1, min(50, ws.max_row + 1)):
        labeled_cols = [c for c in range(1, min(ws.max_column + 1, 60))
                        if is_label(ws.cell(row=r, column=c).value)]
        if len(labeled_cols) >= 3:
            max_label_col = max(labeled_cols)
            for dr in range(r + 1, min(r + 10, ws.max_row + 1)):
                dr_vals = [ws.cell(row=dr, column=c).value
                           for c in range(1, max_label_col + 1)]
                meaningful = [v for v in dr_vals
                              if v is not None and str(v).strip() not in ('', '   ')]
                if len(meaningful) >= 2:
                    return dr, max_label_col
            return r + 1, max_label_col
    return 2, ws.max_column


def is_generic_data_row(vals):
    """True if row has >= 2 non-empty cells and no sum/total marker."""
    meaningful = [v for v in vals
                  if v is not None and str(v).strip() not in ('', '   ')]
    if len(meaningful) < 2:
        return False
    return not any(str(v).strip() in ('Suma:', 'Sum:', 'Total:') for v in vals)


def get_generic_row_key(vals):
    """First non-empty, non-whitespace string value from the row."""
    for v in vals:
        if v is None:
            continue
        s = str(v).strip()
        if s and s not in ('', '   '):
            return s
    return None


def build_old_map(old_rows, is_ls=False, is_lb=False):
    """
    Returns dict: key -> list of vals (ordered by appearance).
    For LS: key = (pozycja, 'parent'|'child').
    For LB: key = pozycja string (any col-A value), duplicates stored in order.
    For others: key = pozycja string.
    """
    result = {}
    for _excel_row, vals in old_rows:
        if is_lb:
            if not is_lb_data_row(vals):
                continue
            key = str(vals[0]).strip()
        else:
            if not is_pozycja(vals[0]):
                continue
            key = str(vals[0]).strip()
            if is_ls:
                key = (key, 'parent' if is_parent_ls(vals) else 'child')
        if key not in result:
            result[key] = []
        result[key].append(vals)
    return result


def build_ls_child_to_parent(rows):
    child_to_parent = {}
    current_parent = None
    for _r, vals in rows:
        if not is_pozycja(vals[0]):
            continue
        if is_parent_ls(vals):
            current_parent = str(vals[0]).strip()
        elif current_parent:
            child_to_parent[str(vals[0]).strip()] = current_parent
    return child_to_parent


def apply_yellow_row(ws, excel_row, max_col):
    for c in range(1, max_col + 1):
        ws.cell(row=excel_row, column=c).fill = YELLOW_FILL


def apply_red_text(ws, excel_row, col_1based):
    if col_1based > MAX_COL:
        return
    cell = ws.cell(row=excel_row, column=col_1based)
    f = cell.font
    cell.font = Font(
        name=f.name, size=f.size, bold=True, italic=f.italic,
        underline=f.underline, strike=f.strike, vertAlign=f.vertAlign,
        color=RED_COLOR
    )


def find_deleted(old_rows, new_rows, is_ls=False, is_lb=False):
    """Return vals of rows present in old but absent (or fewer) in new, in old order."""
    new_occ = {}
    for _, new_vals in new_rows:
        if is_lb:
            if not is_lb_data_row(new_vals):
                continue
            key = str(new_vals[0]).strip()
        else:
            if not is_pozycja(new_vals[0]):
                continue
            pozycja = str(new_vals[0]).strip()
            key = (pozycja, 'parent' if is_parent_ls(new_vals) else 'child') if is_ls else pozycja
        new_occ[key] = new_occ.get(key, 0) + 1

    old_seen = {}
    deleted = []
    for _, vals in old_rows:
        if is_lb:
            if not is_lb_data_row(vals):
                continue
            key = str(vals[0]).strip()
        else:
            if not is_pozycja(vals[0]):
                continue
            pozycja = str(vals[0]).strip()
            key = (pozycja, 'parent' if is_parent_ls(vals) else 'child') if is_ls else pozycja
        old_seen[key] = old_seen.get(key, 0) + 1
        if old_seen[key] > new_occ.get(key, 0):
            deleted.append(vals)
    return deleted


def append_deleted_rows(ws_out, deleted, n_cols=None):
    if not deleted:
        return
    effective_cols = n_cols if n_cols is not None else MAX_COL
    ws_out.append([])
    ws_out.append(['DELETED ELEMENTS:'])
    hdr = ws_out.max_row
    for c in range(1, effective_cols + 1):
        cell = ws_out.cell(row=hdr, column=c)
        cell.fill = DELETE_FILL
        f = cell.font
        cell.font = Font(name=f.name, size=f.size, bold=True, color=RED_COLOR)

    for vals in deleted:
        ws_out.append(vals)
        r = ws_out.max_row
        for c in range(1, effective_cols + 1):
            cell = ws_out.cell(row=r, column=c)
            f = cell.font
            cell.font = Font(name=f.name, size=f.size, strike=True)


def find_generic_deleted(old_data, new_data):
    """Return vals of rows present in old but absent (or fewer) in new, using generic row key.
    Only considers rows with a non-empty first column (position column) to skip summary rows."""
    new_occ = {}
    for _, vals in new_data:
        if not vals or vals[0] is None or str(vals[0]).strip() == '':
            continue
        key = get_generic_row_key(vals)
        if key is not None:
            new_occ[key] = new_occ.get(key, 0) + 1

    old_seen = {}
    deleted = []
    for _, vals in old_data:
        if not vals or vals[0] is None or str(vals[0]).strip() == '':
            continue
        key = get_generic_row_key(vals)
        if key is None:
            continue
        old_seen[key] = old_seen.get(key, 0) + 1
        if old_seen[key] > new_occ.get(key, 0):
            deleted.append(vals)
    return deleted


def compare_sheet(ws_out, old_rows, new_rows, is_ls=False, is_lb=False):
    old_map = build_old_map(old_rows, is_ls=is_ls, is_lb=is_lb)
    # occurrence counter: for duplicate keys, match nth new to nth old
    occurrence = {}
    changed_positions = set()
    row_details = {}  # excel_row -> (old_vals, set of 1-based changed cols)

    for excel_row, new_vals in new_rows:
        if is_lb:
            if not is_lb_data_row(new_vals):
                continue
            pozycja = str(new_vals[0]).strip()
            key = pozycja
        else:
            if not is_pozycja(new_vals[0]):
                continue
            pozycja = str(new_vals[0]).strip()
            parent = is_ls and is_parent_ls(new_vals)
            key = (pozycja, 'parent' if parent else 'child') if is_ls else pozycja

        idx = occurrence.get(key, 0)
        occurrence[key] = idx + 1

        if key not in old_map or idx >= len(old_map[key]):
            apply_yellow_row(ws_out, excel_row, MAX_COL)
            changed_positions.add(pozycja)
            continue

        old_vals = old_map[key][idx]
        n_cols = max(len(new_vals), len(old_vals))
        changed_cols = []
        for i in range(n_cols):
            nv = new_vals[i] if i < len(new_vals) else None
            ov = old_vals[i] if i < len(old_vals) else None
            if not vals_equal(nv, ov):
                changed_cols.append(i + 1)

        if changed_cols:
            apply_yellow_row(ws_out, excel_row, MAX_COL)
            for col in changed_cols:
                apply_red_text(ws_out, excel_row, col)
            changed_positions.add(pozycja)
            row_details[excel_row] = (old_vals, set(changed_cols))

    return changed_positions, row_details


def mark_ls_parents(ws_out, new_rows, changed_children, child_to_parent):
    parents_to_mark = {child_to_parent[c] for c in changed_children if c in child_to_parent}
    if not parents_to_mark:
        return
    for excel_row, vals in new_rows:
        if is_pozycja(vals[0]) and is_parent_ls(vals):
            if str(vals[0]).strip() in parents_to_mark:
                apply_yellow_row(ws_out, excel_row, MAX_COL)


def write_old_values(ws_out, row_details):
    """Next to each changed row, paste old values starting at column M with red on changed cells."""
    for excel_row, (old_vals, changed_cols) in row_details.items():
        for i, val in enumerate(old_vals):
            if isinstance(val, float):
                val = round(val, 2)
            cell = ws_out.cell(row=excel_row, column=OLD_COL_START + i)
            cell.value = val
            if (i + 1) in changed_cols:
                f = cell.font
                cell.font = Font(name=f.name, size=f.size, bold=True, color=RED_COLOR)


def write_generic_old_values(ws_out, row_details, max_data_col):
    """Write old values starting at column M for each changed row (generic mode)."""
    for excel_row, (old_vals, changed_cols) in row_details.items():
        for i in range(max_data_col):
            val = old_vals[i] if i < len(old_vals) else None
            if isinstance(val, float):
                val = round(val, 2)
            cell = ws_out.cell(row=excel_row, column=OLD_COL_START + i)
            cell.value = val
            if (i + 1) in changed_cols:
                f = cell.font
                cell.font = Font(name=f.name, size=f.size, bold=True, color=RED_COLOR)
    for i in range(max_data_col):
        src = get_column_letter(i + 1)
        dst = get_column_letter(OLD_COL_START + i)
        if src in ws_out.column_dimensions:
            ws_out.column_dimensions[dst].width = ws_out.column_dimensions[src].width


def clear_print_area_xml(xlsx_path):
    """
    Remove all Print_Area defined names directly from workbook.xml inside the xlsx zip.
    openpyxl re-generates these on save, so we must strip them post-save.
    """
    ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

    tmp_path = xlsx_path + '.tmp'
    with zipfile.ZipFile(xlsx_path, 'r') as zin, \
         zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/workbook.xml':
                tree = ET.fromstring(data)
                defined_names = tree.find(f'{{{NS}}}definedNames')
                if defined_names is not None:
                    to_remove = [
                        dn for dn in defined_names
                        if 'Print_Area' in dn.get('name', '')
                    ]
                    for dn in to_remove:
                        defined_names.remove(dn)
                    # Remove empty definedNames element too
                    if len(defined_names) == 0:
                        tree.remove(defined_names)
                data = ET.tostring(tree, encoding='UTF-8', xml_declaration=True)
            zout.writestr(item, data)

    os.replace(tmp_path, xlsx_path)


def clean_comparison_file(file_path, log_cb):
    """Remove old-values columns (M+) from sheets that have them; set print area for ALL sheets."""
    wb = load_workbook(file_path)

    # Detect sheets that have old values in columns M–AB
    sheets_with_old = set(
        ws.title for ws in wb.worksheets
        if any(ws.cell(row=r, column=c).value is not None
               for r in range(1, min(ws.max_row + 1, 500))
               for c in range(OLD_COL_START, OLD_COL_START + 15))
    )
    # Fallback for standard sheet names
    if not sheets_with_old:
        sheets_with_old = set(s for s in DATA_SHEETS if s in wb.sheetnames)

    if sheets_with_old:
        log_cb(f'Czyszczenie starych wartości: {", ".join(sorted(sheets_with_old))}')
    else:
        log_cb('Brak starych wartości — ustawiam tylko obszary wydruku.')

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        # Clear old values only from sheets that have them
        if sheet in sheets_with_old:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.column >= OLD_COL_START:
                        cell.value = None
                        cell.fill = PatternFill()
                        cell.font = Font()
            for col_idx in range(OLD_COL_START, OLD_COL_START + 60):
                col_letter = get_column_letter(col_idx)
                if col_letter in ws.column_dimensions:
                    del ws.column_dimensions[col_letter]

        # Find rightmost non-empty column (capped at col L) and last data row
        # by scanning actual content — works for all sheet structures.
        right_col = 1
        last_row = 1
        cap = min(OLD_COL_START - 1, ws.max_column)
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == 'DELETED ELEMENTS:':
                break
            for c in range(cap, 0, -1):
                if ws.cell(row=r, column=c).value is not None:
                    if c > right_col:
                        right_col = c
                    last_row = r
                    break

        ws.print_area = f'A1:{get_column_letter(right_col)}{last_row}'
        log_cb(f'  {sheet}: obszar wydruku A1:{get_column_letter(right_col)}{last_row}')

    log_cb('Zapisywanie...')
    wb.save(file_path)
    log_cb('Gotowe!')


def run_comparison(old_file, new_file, data_start, log_cb, lg_data_start=None):
    if lg_data_start is None:
        lg_data_start = data_start

    out_dir = os.path.dirname(os.path.abspath(new_file))
    base_name = os.path.splitext(os.path.basename(new_file))[0]
    out_file = os.path.join(out_dir, base_name + '_POROWNANIE.xlsx')

    log_cb(f'Kopiowanie nowego pliku → {os.path.basename(out_file)}')
    shutil.copy2(new_file, out_file)

    wb_old_d = load_workbook(old_file, data_only=True)
    wb_new_d = load_workbook(new_file, data_only=True)
    wb_out = load_workbook(out_file)

    total_changed = 0
    for sheet in DATA_SHEETS:
        if sheet not in wb_old_d.sheetnames or sheet not in wb_out.sheetnames:
            log_cb(f'  {sheet}: pominięty (brak w jednym z plików)')
            continue

        sheet_start = lg_data_start if sheet == 'LG' else data_start
        old_rows = read_data_rows(wb_old_d, sheet, sheet_start)
        new_rows = read_data_rows(wb_new_d, sheet, sheet_start)
        ws_out = wb_out[sheet]

        is_ls = sheet == 'LS'
        is_lb = sheet == 'LB'
        changed, row_details = compare_sheet(ws_out, old_rows, new_rows, is_ls=is_ls, is_lb=is_lb)
        write_old_values(ws_out, row_details)
        for i in range(1, MAX_COL + 1):
            ws_out.column_dimensions[get_column_letter(OLD_COL_START + i - 1)].width = \
                ws_out.column_dimensions[get_column_letter(i)].width

        if is_ls:
            child_to_parent = build_ls_child_to_parent(new_rows)
            changed_children = set()
            for _r, vals in new_rows:
                if is_pozycja(vals[0]) and not is_parent_ls(vals):
                    if str(vals[0]).strip() in changed:
                        changed_children.add(str(vals[0]).strip())
            mark_ls_parents(ws_out, new_rows, changed_children, child_to_parent)

        deleted = find_deleted(old_rows, new_rows, is_ls=is_ls, is_lb=is_lb)
        append_deleted_rows(ws_out, deleted)

        log_cb(f'  {sheet}: {len(changed)} zmienione, {len(deleted)} usunięte')
        total_changed += len(changed)

    log_cb('Zapisywanie...')
    wb_out.save(out_file)

    log_cb('Czyszczenie obszarów wydruku...')
    clear_print_area_xml(out_file)

    log_cb(f'\nGotowe! Łącznie zmian: {total_changed}')
    log_cb(f'Plik: {out_file}')
    return out_file


def run_generic_comparison(old_file, new_file, log_cb, excluded_sheets=None):
    """
    Compare all sheets present in both files (minus excluded_sheets).
    Auto-detects header row and data start for each sheet.
    Writes old values for changed rows starting at column M.
    """
    if excluded_sheets is None:
        excluded_sheets = set()

    out_dir = os.path.dirname(os.path.abspath(new_file))
    base_name = os.path.splitext(os.path.basename(new_file))[0]
    out_file = os.path.join(out_dir, base_name + '_POROWNANIE.xlsx')

    log_cb(f'Kopiowanie nowego pliku → {os.path.basename(out_file)}')
    shutil.copy2(new_file, out_file)

    wb_old_d = load_workbook(old_file, data_only=True)
    wb_new_d = load_workbook(new_file, data_only=True)
    wb_out = load_workbook(out_file)

    old_sheet_set = set(wb_old_d.sheetnames)
    sheets_to_compare = [s for s in wb_out.sheetnames
                         if s in old_sheet_set and s not in excluded_sheets]
    skipped = [s for s in wb_out.sheetnames if s in old_sheet_set and s in excluded_sheets]
    log_cb(f'Porównywane ({len(sheets_to_compare)}): {", ".join(sheets_to_compare)}')
    if skipped:
        log_cb(f'Pominięte: {", ".join(skipped)}')

    total_changed = 0
    for sheet in sheets_to_compare:
        ws_old = wb_old_d[sheet]
        ws_new = wb_new_d[sheet]
        ws_out_s = wb_out[sheet]

        data_start_old, _ = detect_data_start(ws_old)
        data_start_new, max_data_col = detect_data_start(ws_new)
        log_cb(f'  {sheet}: wiersze od {data_start_new} (nowy), {data_start_old} (stary), '
               f'kolumny 1–{max_data_col}')

        old_rows = [(r, [ws_old.cell(row=r, column=c).value
                         for c in range(1, ws_old.max_column + 1)])
                    for r in range(data_start_old, ws_old.max_row + 1)]
        new_rows = [(r, [ws_new.cell(row=r, column=c).value
                         for c in range(1, ws_new.max_column + 1)])
                    for r in range(data_start_new, ws_new.max_row + 1)]

        old_data = [(r, v) for r, v in old_rows if is_generic_data_row(v)]
        new_data = [(r, v) for r, v in new_rows if is_generic_data_row(v)]

        old_map = {}
        for _, vals in old_data:
            key = get_generic_row_key(vals)
            if key is not None:
                old_map.setdefault(key, []).append(vals)

        occurrence = {}
        changed_count = 0
        row_details = {}

        for excel_row, new_vals in new_data:
            key = get_generic_row_key(new_vals)
            if key is None:
                continue

            idx = occurrence.get(key, 0)
            occurrence[key] = idx + 1

            if key not in old_map or idx >= len(old_map[key]):
                apply_yellow_row(ws_out_s, excel_row, max_data_col)
                changed_count += 1
                continue

            old_vals = old_map[key][idx]
            changed_cols = []
            for i in range(max_data_col):
                nv = new_vals[i] if i < len(new_vals) else None
                ov = old_vals[i] if i < len(old_vals) else None
                if not vals_equal(nv, ov):
                    changed_cols.append(i + 1)

            if changed_cols:
                apply_yellow_row(ws_out_s, excel_row, max_data_col)
                for col in changed_cols:
                    cell = ws_out_s.cell(row=excel_row, column=col)
                    f = cell.font
                    cell.font = Font(
                        name=f.name, size=f.size, bold=True, italic=f.italic,
                        underline=f.underline, strike=f.strike,
                        vertAlign=f.vertAlign, color=RED_COLOR
                    )
                row_details[excel_row] = (old_vals, set(changed_cols))
                changed_count += 1

        write_generic_old_values(ws_out_s, row_details, max_data_col)

        deleted = find_generic_deleted(old_data, new_data)
        append_deleted_rows(ws_out_s, deleted, n_cols=max_data_col)

        log_cb(f'  {sheet}: {changed_count} zmienione, {len(deleted)} usunięte')
        total_changed += changed_count

    log_cb('Zapisywanie...')
    wb_out.save(out_file)

    log_cb('Czyszczenie starych obszarów wydruku...')
    clear_print_area_xml(out_file)

    log_cb(f'\nGotowe! Łącznie zmian: {total_changed}')
    log_cb(f'Plik: {out_file}')
    return out_file


# ── GUI ───────────────────────────────────────────────────────────────────────

class App(TkinterDnD.Tk):
    def __init__(self):
        super().__init__()
        self.title('BOM Comparer')
        self.resizable(False, False)
        self.configure(padx=20, pady=16)

        self.old_path = tk.StringVar()
        self.new_path = tk.StringVar()
        self.clean_path = tk.StringVar()
        self.sheet_vars = {}   # sheet_name -> BooleanVar (True = compare)

        self._build_ui()

    def _build_ui(self):
        HINT = 'Przeciągnij plik lub kliknij aby wybrać ścieżkę'

        # ── File drop zones ──
        for label_text, var, attr, frame_attr in [
            ('Stara lista', self.old_path, 'old_drop', None),
            ('Nowa lista',  self.new_path, 'new_drop', 'new_drop_frame'),
        ]:
            frame = tk.LabelFrame(self, text=label_text, padx=8, pady=8)
            frame.pack(fill='x', pady=4)
            if frame_attr:
                setattr(self, frame_attr, frame)

            drop = tk.Label(
                frame, text=HINT,
                relief='groove', bg='#f0f0f0',
                width=56, height=3,
                wraplength=420, justify='center', anchor='center'
            )
            drop.pack(side='left', fill='x', expand=True, padx=(0, 6))

            clear_btn = tk.Button(
                frame, text='✕', width=3,
                fg='#c62828', relief='flat', cursor='hand2',
                command=lambda v=var, d=drop, h=HINT: self._clear_path(v, d, h)
            )
            clear_btn.pack(side='left')

            drop.drop_target_register(DND_FILES)
            drop.dnd_bind('<<Drop>>', lambda e, v=var, d=drop: self._on_drop(e, v, d))
            drop.bind('<Button-1>', lambda e, v=var, d=drop: self._browse(v, d))
            setattr(self, attr, drop)

        # ── Sheet selection (shown after both files are loaded) ──
        self.sheets_frame = tk.LabelFrame(
            self, text='Pomiń arkusze  (odznacz, aby wykluczyć z porównania)',
            padx=8, pady=6
        )
        # Not packed yet — appears dynamically in _update_sheets()

        # ── Run button ──
        self.run_btn = tk.Button(self, text='▶  Porównaj', command=self._run,
                                 bg='#2e7d32', fg='white', font=('Arial', 11, 'bold'),
                                 padx=16, pady=6, relief='flat', cursor='hand2')
        self.run_btn.pack(pady=10)

        # ── Log output ──
        log_frame = tk.LabelFrame(self, text='Postęp', padx=8, pady=8)
        log_frame.pack(fill='both', expand=True, pady=4)
        self.log = tk.Text(log_frame, height=10, state='disabled',
                           font=('Consolas', 9), bg='#1e1e1e', fg='#d4d4d4',
                           relief='flat', wrap='word')
        self.log.pack(fill='both', expand=True)
        sb = ttk.Scrollbar(log_frame, command=self.log.yview)
        self.log['yscrollcommand'] = sb.set

        # ── Clean section ──
        CLEAN_HINT = 'Przeciągnij plik z porównaniem lub kliknij aby wybrać ścieżkę'
        clean_frame = tk.LabelFrame(self, text='Przygotuj do druku', padx=8, pady=8)
        clean_frame.pack(fill='x', pady=(8, 4))

        self.clean_drop = tk.Label(
            clean_frame, text=CLEAN_HINT,
            relief='groove', bg='#f0f0f0',
            width=56, height=3,
            wraplength=420, justify='center', anchor='center'
        )
        self.clean_drop.pack(side='left', fill='x', expand=True, padx=(0, 6))

        tk.Button(
            clean_frame, text='✕', width=3,
            fg='#c62828', relief='flat', cursor='hand2',
            command=lambda: self._clear_path(self.clean_path, self.clean_drop, CLEAN_HINT)
        ).pack(side='left')

        self.clean_drop.drop_target_register(DND_FILES)
        self.clean_drop.dnd_bind('<<Drop>>', lambda e: self._on_drop(e, self.clean_path, self.clean_drop))
        self.clean_drop.bind('<Button-1>', lambda e: self._browse(self.clean_path, self.clean_drop))

        self.clean_btn = tk.Button(
            self, text='🖨  Usuń stare wartości i ustaw obszar wydruku',
            command=self._clean,
            bg='#1565c0', fg='white', font=('Arial', 10, 'bold'),
            padx=12, pady=5, relief='flat', cursor='hand2'
        )
        self.clean_btn.pack(pady=(4, 8))

    def _update_sheets(self):
        """Refresh sheet checkbox list after a file is loaded or cleared."""
        for w in self.sheets_frame.winfo_children():
            w.destroy()
        self.sheet_vars = {}

        old = self.old_path.get()
        new = self.new_path.get()
        if not old or not new or not os.path.isfile(old) or not os.path.isfile(new):
            self.sheets_frame.pack_forget()
            return

        try:
            wb_old = load_workbook(old, read_only=True)
            wb_new = load_workbook(new, read_only=True)
            old_names = set(wb_old.sheetnames)
            common = [s for s in wb_new.sheetnames if s in old_names]
            wb_old.close()
            wb_new.close()
        except Exception:
            self.sheets_frame.pack_forget()
            return

        if not common:
            self.sheets_frame.pack_forget()
            return

        cols = 3
        for i, sheet in enumerate(common):
            var = tk.BooleanVar(value=True)   # True = include in comparison
            self.sheet_vars[sheet] = var
            tk.Checkbutton(
                self.sheets_frame, text=sheet, variable=var,
                anchor='w', wraplength=160
            ).grid(row=i // cols, column=i % cols, sticky='w', padx=6, pady=2)

        self.sheets_frame.pack(fill='x', pady=4, after=self.new_drop_frame)

    def _clear_path(self, var, label, hint):
        var.set('')
        label.config(text=hint, bg='#f0f0f0')
        self._update_sheets()

    def _on_drop(self, event, var, label):
        path = event.data.strip().strip('{}')
        if path.lower().endswith('.xlsx'):
            var.set(path)
            label.config(text=os.path.basename(path) + '\n' + path, bg='#e8f5e9')
            self._update_sheets()
        else:
            messagebox.showerror('Błąd', 'Proszę przeciągnąć plik .xlsx')

    def _browse(self, var, label):
        path = filedialog.askopenfilename(
            filetypes=[('Excel files', '*.xlsx')],
            title='Wybierz plik Excel'
        )
        if path:
            var.set(path)
            label.config(text=os.path.basename(path) + '\n' + path, bg='#e8f5e9')
            self._update_sheets()

    def _log(self, msg):
        self.log.config(state='normal')
        self.log.insert('end', msg + '\n')
        self.log.see('end')
        self.log.config(state='disabled')

    def _run(self):
        old = self.old_path.get()
        new = self.new_path.get()
        if not old or not os.path.isfile(old):
            messagebox.showerror('Błąd', 'Wybierz starą listę materiałową.')
            return
        if not new or not os.path.isfile(new):
            messagebox.showerror('Błąd', 'Wybierz nową listę materiałową.')
            return

        excluded = {s for s, v in self.sheet_vars.items() if not v.get()}

        self.run_btn.config(state='disabled', text='Przetwarzanie...')
        self.log.config(state='normal')
        self.log.delete('1.0', 'end')
        self.log.config(state='disabled')

        def task():
            try:
                run_generic_comparison(old, new, self._log, excluded_sheets=excluded)
            except Exception as ex:
                self._log(f'\n❌ Błąd: {ex}')
            finally:
                self.run_btn.config(state='normal', text='▶  Porównaj')

        threading.Thread(target=task, daemon=True).start()

    def _clean(self):
        path = self.clean_path.get()
        if not path or not os.path.isfile(path):
            messagebox.showerror('Błąd', 'Wybierz plik z porównaniem (_POROWNANIE.xlsx).')
            return

        self.clean_btn.config(state='disabled', text='Przetwarzanie...')
        self.log.config(state='normal')
        self.log.delete('1.0', 'end')
        self.log.config(state='disabled')

        def task():
            try:
                clean_comparison_file(path, self._log)
            except Exception as ex:
                self._log(f'\n❌ Błąd: {ex}')
            finally:
                self.clean_btn.config(
                    state='normal',
                    text='🖨  Usuń stare wartości i ustaw obszar wydruku'
                )

        threading.Thread(target=task, daemon=True).start()


if __name__ == '__main__':
    app = App()
    app.mainloop()
